import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

FILE_PATH = r"C:\Users\aphiwut.j\Documents\Learning\resulting processing\test\merged_final_scores_2.csv"  # your CSV file

# --- STEP 1: Load CSV ---
input_csv = FILE_PATH
df = pd.read_csv(input_csv)

# Dropping "Section" and "SIS Login ID"
columns_to_drop = ['Section', 'SIS Login ID']
df = df.drop(columns_to_drop, axis=1)

# --- STEP 2: Rename columns ---
df = df.rename(columns={
    "Enrol No": "Enrolment No",
    "SIS User ID": "Student Number"
})

# --- STEP 3: Identify columns ---
id_vars = ["Enrolment No", "Student Number"]
score_cols = [col for col in df.columns if col not in id_vars]

# --- STEP 4: Unpivot ---
df_long = df.melt(
    id_vars=id_vars,
    value_vars=score_cols,
    var_name="Score Type",
    value_name="Outcome"
)

def convert_outcome(x):
    
    try:
        tmp_x = float(x)
        if tmp_x == 100:
            return "Competency achieved/Pass"
        else:
            return "Competency not achieved/Fail"
    except Exception as e:
        pass
    # if isinstance(x, (int) or isinstance(x, float)):
    #     print("int or float")
    #     if x == 100:
    #         return "Competency achieved/Pass"
    #     elif x < 100:
    #         return "Competency not achieved/Fail"
    if pd.isna(x):
        # print("nan")
        return "Record not found"
    if isinstance(x, str) and x.lower() == "ex": #change to startswith
        # print("Ex")
        return "Credit Transfer/national recognition"
    else:
        return "Invalid input"


df_long["Outcome"] = df_long["Outcome"].apply(convert_outcome)

# --- STEP 6: Save to Excel ---
output_file = "output.xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Original", index=False)
    df_long.to_excel(writer, sheet_name="Unpivoted", index=False)

# --- STEP 7: Apply colours in Excel ---
wb = load_workbook(output_file)
ws = wb["Unpivoted"]

# Colour definitions (HEX RGB)
green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
red_fill   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
yellow_fill   = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

# Find the Outcome column index
outcome_col = None
for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == "Outcome":
        outcome_col = col
        break

# Apply colours row by row
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=outcome_col)
    if cell.value == "Competency achieved/Pass":
        cell.fill = green_fill
    elif cell.value == "Credit Transfer/national recognition":
        cell.fill = yellow_fill
    elif cell.value == "Competency not achieved/Fail":
        cell.fill = red_fill
    else:
        cell.fill = black_fill

wb.save(output_file)

print("File saved with colour formatting applied.")
